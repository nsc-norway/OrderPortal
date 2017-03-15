"OrderPortal: File check processors."

from __future__ import print_function, absolute_import

from .baseprocessor import BaseProcessor
import openpyxl
import io

output_table_header_prefixes = [
        'Plate',
        'Position',
        'Sample Name',
        'Conc',
        'A260/280',
        'A260/230',
        'Volume provided',
        'Total DNA / RNA',
        'Index name',
        'Index Seq',
        'Primers, Linkers',
        'Approx no. Reads',
        #'Comments'
        ]
MAX_SAMPLES = 2000


class SampleTableImporter(BaseProcessor):
    "The value must be a text file."

    def run(self, value, **kwargs):
        file_content = kwargs.get('body')
        if file_content:
            try:
                try:
                    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
                except openpyxl.utils.exceptions.InvalidFileException:
                    raise ValueError("Unable to read the file, please make sure it is in Excel (xlsx) format")
                ws = wb.worksheets[0]
                MAX_COLUMNS=20
                headers = [
                        (str(ws.cell(column=i, row=1).value), i)
                        for i in range(1,20)
                        if ws.cell(column=i,row=1).data_type != 'n' # Null
                        ]
                input_index_for_outputs = []
                for i, header_prefix in enumerate(output_table_header_prefixes):
                    for header, input_pos in headers:
                        if header.startswith("Sample Name"):
                            sample_name_index = input_pos
                        elif header.startswith("Number"):
                            sample_number_index = input_pos
                        if header.startswith(header_prefix):
                            input_index_for_outputs.append(input_pos)
                            break
                    else:
                        raise ValueError("Missing column " + header_prefix)

                test = ws.cell(row=2, column=sample_number_index).value
                if str(test) == "0": # Sample number 0 is example
                    start_row = 3
                else:
                    start_row = 2

                list_data = []
                for i in range(start_row, MAX_SAMPLES):
                    name = ws.cell(row=i, column=sample_name_index).value
                    if name:
                        row_data = []
                        for input_index in input_index_for_outputs:
                            val = ws.cell(row=i, column=input_index).value
                            row_data.append(val if val is not None else "")
                        list_data.append(row_data)
                    else:
                        break
                self.order['fields']['samples'] = list_data
            except UnicodeEncodeError:
                raise ValueError("Unsupported character in sample file. Please use only A-Z and numbers.")

