import itertools
import csv
import io
import traceback # TODO remove
from openpyxl import load_workbook

class ValidationError(Exception):
    """Used internally for validation functions"""
    pass

def int_val(v):
    if isinstance(v, int) or str(v).isdigit():
        return int(v)
    else:
        raise ValidationError("Enter a number")

str_val = str

def float_val(v):
    try:
        return float(str(v).replace(",","."))
    except ValueError:
        raise ValidationError("Enter a decimal number")

def index_seq(v):
    if set(str(v).upper()) <= set("AGCT-"):
        return str(v).upper()
    else:
        raise ValidationError("Use only A,C,G,T and hyphen")


class Field(object):
    def __init__(self, id, label, required, validator, width):
        self.id = id
        self.label = label
        self.required = required
        self.validator = validator
        self.width = width

# First arg:  Internal label
# Second arg: Header in spreadsheet / CSV / table
SAMPLE_FIELDS = [
        Field("plate",               "Plate",            False, str_val,       4),
        Field("position",            "Position",         False, str_val,       2),
        Field("sample_name",         "Sample name",      True,  str_val,       10),
        Field("conc",                "Conc.",            True,  float_val,     4),
        Field("a_260_280",           "A260/280",         False, float_val,     4),
        Field("a_260_230",           "A260/230",         False, float_val,     4),
        Field("volume",              "Volume provided",  False, float_val,     4),
        Field("total_dna_rna",       "Total DNA / RNA",  False, float_val,     4),
        Field("index_name",          "Index name",       False, str_val,       8),
        Field("index_seq",           "Index Seq",        False, index_seq,     15),
        Field("primers",             "Primers, Linkers or RE sites present?",   False, str_val, 10),
        Field("num_reads",           "Approx no. Reads, Gb or lanes requested", False, str_val,  8)
        ]

MAX_SAMPLES = 4 # TODO 16000

class ImportException(Exception):
    pass


def import_excel_file(buffer):
    """Simple importer which reads the sample table Excel file and
    produces a list of dicts, one for each sample (identical to the
    format stored in the DB and exported to the LIMS, etc.)

    Only minimal validation is done. On error, an ImportException
    is raised, with a message indicating the cause.
    """
    fh = io.BytesIO(buffer)
    try:
        wb = load_workbook(fh)
    except:
        traceback.print_exc()
        raise ImportException("The file is not a valid Excel (xlsx) spreadsheet")
    try:
        ws = wb.worksheets[0]
    except IndexError:
        raise ImportException("No worksheets found in the file")

    col_of = {}

    for i in range(20):
        val = ws.cell(row=1, column=i+1).value
        if val:
            for field in SAMPLE_FIELDS:
                if unicode(val).lower().startswith(field.label.lower()):
                    col_of[field.id] = i+1

    if not col_of.has_key('sample_name'):
        raise ImportException("Missing column for Sample Name")
    samples = []
    for i in range(3, 3+MAX_SAMPLES):
        name = ws.cell(row=i, column=col_of['sample_name']).value
        if name:
            sample = {}
            for field in SAMPLE_FIELDS:
                if col_of.has_key(field.id):
                    val = ws.cell(row=i, column=col_of[field.id]).value
                    if val is not None:
                        sample[field.id] = val
            samples.append(sample)
        else:
            break

    if ws.cell(row=i+1, column=col_of['sample_name']).value:
        raise ImportException(("We only support importing {0} samples. Please "+
                "contact NSC staff if you would like to submit more samples."
                ).format(MAX_SAMPLES))

    return samples

def import_csv_file(buffer):
    fh = io.BytesIO(buffer)
    try:
        reader = csv.DictReader(fh)
        samples = list(reader)
    except e:
        raise ImportException("Failed to read the CSV format (" +str(e)+ ")")
    if any(not sample.get('sample_name') for sample in samples):
        raise ImportException("Missing sample name")
    return samples

class Cell(object):
    def __init__(self, field, valid, value, message):
        self.field = field
        self.valid = valid
        self.value = value
        self.message = message


def validate_table(samples_raw):
    """Main sample table validation.

    Takes a preliminary grid, from either file import or direct user import.
    Input format: list of dict objects, each dict object representing one
    sample, containing the sample field names as keys, and values as values.

    Returns:
        (validation_table, sample_list)

    validation_table: A list containing feedback to the user. Contains a
    list of lists of Cell objects. The first dimension is the row, nested
    dimension is the column.

    sample_list: list similar to samples_raw, but only contains valid data
    (input cells which fail validation are not included). All data in
    sample_list have the specified data types.
    """

    validation_table = []
    sample_list = []
    for sample_raw in samples_raw:
        sample = {}
        row = []
        for field in SAMPLE_FIELDS:
            value = sample_raw.get(field.id, '')
            try:
                if field.required:
                    if value == "":
                        raise ValidationError("Missing value")
                if value != "":
                    value = field.validator(value)
                if value is not None:
                    sample[field.id] = value
                row.append(Cell(field, True, value, None))
            except ValidationError, e:
                row.append(Cell(field, False, value, str(e)))
        sample_list.append(sample)
        validation_table.append(row)
    return (validation_table, sample_list)
